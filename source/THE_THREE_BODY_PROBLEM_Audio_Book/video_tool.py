import os
from moviepy.editor import VideoFileClip
from openpyxl import load_workbook


def extract_audio_segment(video_path, start_time, end_time, output_path=None):
    """
    从视频中提取指定时间段的音频

    参数:
    video_path: 视频文件路径
    start_time: 开始时间(秒)
    end_time: 结束时间(秒)
    output_path: 输出音频文件路径(可选)
    """
    try:
        # 如果没有指定输出路径，使用默认名称
        if output_path is None:
            name = os.path.splitext(video_path)[0]
            output_path = f"{name}_audio_{start_time}s_to_{end_time}s.mp3"

        # 加载视频并提取指定时间段的音频
        with VideoFileClip(video_path) as video:
            # 提取指定时间段的音频
            audio_segment = video.audio.subclip(start_time, end_time)

            # 保存音频文件
            audio_segment.write_audiofile(output_path)

        print(f"音频已成功提取到: {output_path}")
        return True

    except Exception as e:
        print(f"处理过程中出现错误: {str(e)}")
        return False


# 将"小时:分:秒"格式的时间转换为秒数
def time_to_seconds(time_str):
    # 分割时间字符串
    parts = time_str.split(':')

    # 根据时间格式可能有不同的部分数量
    if len(parts) == 3:  # 小时:分钟:秒
        hours, minutes, seconds = parts
        return int(hours) * 3600 + int(minutes) * 60 + float(seconds)
    elif len(parts) == 2:  # 分钟:秒
        minutes, seconds = parts
        return int(minutes) * 60 + float(seconds)
    else:
        raise ValueError(f"无法解析的时间格式: {time_str}")


# 在这里设置您的参数
if __name__ == "__main__":
    # 视频文件路径 - 修改为您要处理的视频文件路径
    VIDEO_PATH = "D:/Blogs/THE_THREE_BODY_PROBLEM_Audio_Book/1-1疯狂年代_The_Madness_Years英_音三_体_The_Three-Body_Pro.mp4"  # 请替换为实际视频路径

    # 读取Excel文件
    excel_file = "1-1疯狂年代_The_Madness_Years.xlsx"
    wb = load_workbook(excel_file)
    ws = wb.active  # 获取活动工作表

    for n in range(9, 10):
        # 设置要读取的行号n（可以根据需要修改）
        # n = 6  # 例如读取第1行

        # 音频提取时间段(秒) - 修改为您需要的时间段
        START_TIME = time_to_seconds(str(ws.cell(row=n, column=1).value))  # 开始时间(秒)
        END_TIME = time_to_seconds(str(ws.cell(row=n, column=2).value))  # 结束时间(秒)

        # 输出文件路径(可选) - 如果为None则自动生成
        OUTPUT_PATH = str(ws.cell(row=n, column=3).value) + '.mp3'  # 可以设置为具体路径，如 "output_audio.mp3"

        # 执行音频提取
        extract_audio_segment(VIDEO_PATH, START_TIME, END_TIME, OUTPUT_PATH)